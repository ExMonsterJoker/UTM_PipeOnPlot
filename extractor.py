import os
import openpyxl
import datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from concurrent.futures import ProcessPoolExecutor, as_completed
import multiprocessing
from functools import lru_cache
import time
from typing import Dict, List, Tuple, Optional, Any
import logging
import argparse  # Import argparse

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# --- Configuration ---
DATA_FOLDER = "Raw Data"
OUTPUT_FOLDER = "Outputs"
MAX_WORKERS = max(1, multiprocessing.cpu_count() - 1)

# Sheet names for different report versions
SHEET_1_NAME_V1 = "TEMPLATE INSPEKSI"
SHEET_1_NAME_V1_ALT = "TEMPALTE INSPEKSI"
SHEET_2_NAME_V1 = "Visual"

SHEET_1_NAME_V2 = "UT Data"
SHEET_2_NAME_V2 = "General Visual"

# Search range for the "REMARKS" cell in both sheets
SEARCH_REMARKS_END_COL = 'Az'
SEARCH_REMARKS_END_ROW = 30

# Global cache for column lookups
_column_cache = {}


# --- Performance Optimized Helper Functions ---

def get_column_values_batch(sheet: Worksheet, col_idx: int, start_row: int, end_row: int) -> List[Any]:
    """Reads all cell values in a column range at once for better performance."""
    if not col_idx or start_row > end_row:
        return []
    return [row[0] for row in
            sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=col_idx, max_col=col_idx, values_only=True)]


def get_area_values_batch(sheet: Worksheet, area: Tuple[int, int, int, int]) -> Dict[Tuple[int, int], Any]:
    """Read all cell values in an area at once for better performance."""
    if not area:
        return {}

    start_row, end_row, start_col, end_col = area
    values = {}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row,
                                                  min_col=start_col, max_col=end_col, values_only=True)):
        for col_idx, value in enumerate(row):
            if value is not None:
                values[(start_row + row_idx, start_col + col_idx)] = value

    return values


def search_for_value_optimized(sheet: Worksheet, area: Tuple[int, int, int, int],
                               keywords: List[str], offset: int = 0, row_offset: int = 0) -> str:
    """Optimized search with early termination and batch reading."""
    if not area:
        return "Not Found"

    area_values = get_area_values_batch(sheet, area)
    keywords_lower = [k.lower() for k in keywords]

    for (row, col), value in area_values.items():
        if value and isinstance(value, str):
            value_lower = value.lower()
            for keyword in keywords_lower:
                if keyword in value_lower:
                    target_row = row + row_offset
                    target_col = col + offset

                    if target_row <= sheet.max_row and target_col <= sheet.max_column:
                        cell_value = sheet.cell(row=target_row, column=target_col).value
                        if isinstance(cell_value, tuple) and len(cell_value) == 1:
                            return cell_value[0]
                        return cell_value

    return "Not Found"


def find_value_in_cell_optimized(sheet: Worksheet, area: Tuple[int, int, int, int], keyword: str) -> str:
    """Optimized version of find_value_in_cell."""
    if not area:
        return "Not Found"

    area_values = get_area_values_batch(sheet, area)
    keyword_lower = keyword.lower()

    for (row, col), value in area_values.items():
        if value and isinstance(value, str) and keyword_lower in value.lower():
            return value

    return "Not Found"


@lru_cache(maxsize=128)
def find_column_by_keyword_cached(sheet_title: str, header_area: Tuple[int, int, int, int], keyword: str) -> Optional[
    int]:
    """Cached version of column finder to avoid repeated searches."""
    return None


def find_column_by_keyword_optimized(sheet: Worksheet, header_area: Tuple[int, int, int, int], keyword: str,
                                     filename: str) -> Optional[int]:
    """Optimized column finder with batch reading."""
    if not header_area:
        return None

    cache_key = (sheet.title, header_area, keyword.lower())
    if cache_key in _column_cache:
        return _column_cache[cache_key]

    area_values = get_area_values_batch(sheet, header_area)
    keyword_lower = keyword.lower()

    for (row, col), value in area_values.items():
        if value and isinstance(value, str) and keyword_lower in value.lower():
            _column_cache[cache_key] = col
            return col

    logger.warning(f"[{filename}] Could not find column for keyword '{keyword}' in sheet '{sheet.title}'")
    _column_cache[cache_key] = None
    return None


def find_check_mark_in_column_optimized(sheet: Worksheet, header_area: Tuple[int, int, int, int],
                                        data_area: Tuple[int, int, int, int], keyword: str, filename: str) -> str:
    """Optimized check mark finder that supports normal check symbol and Wingdings 2 'P'."""
    if not header_area or not data_area:
        return "Not Found"

    target_col = find_column_by_keyword_optimized(sheet, header_area, keyword, filename)
    if not target_col:
        return "Not Found"

    data_start_row, data_end_row, _, _ = data_area
    check_symbols = {"√", "P"}

    column_values = get_column_values_batch(sheet, target_col, data_start_row, data_end_row)

    for cell_val in column_values:
        if cell_val and isinstance(cell_val, str) and cell_val.strip() in check_symbols:
            return "√"

    return "Not Found"


def extract_unique_values_from_col_optimized(sheet: Worksheet, data_area: Tuple[int, int, int, int],
                                             col_idx: int) -> str:
    """Optimized unique value extraction."""
    if not data_area or not col_idx:
        return "Not Found"

    data_start_row, data_end_row, _, _ = data_area
    unique_values = set()

    column_values = get_column_values_batch(sheet, col_idx, data_start_row, data_end_row)

    for cell_val in column_values:
        if cell_val and str(cell_val).strip():
            unique_values.add(str(cell_val).strip())

    return ", ".join(sorted(list(unique_values))) if unique_values else "Not Found"


def extract_unique_values_from_col_or_notfound(sheet: Worksheet, data_area: Tuple[int, int, int, int],
                                               col_idx: int) -> str:
    """Extract unique non-empty values from a column or return 'Not Found' if none."""
    if not data_area or not col_idx:
        return "Not Found"

    data_start_row, data_end_row, _, _ = data_area
    unique_values = set()

    column_values = get_column_values_batch(sheet, col_idx, data_start_row, data_end_row)

    for cell_val in column_values:
        if cell_val and str(cell_val).strip():
            unique_values.add(str(cell_val).strip())
    return ", ".join(sorted(list(unique_values))) if unique_values else "Not Found"


# --- Optimized Area Finding Functions ---

def find_area_by_remarks_optimized(sheet: Worksheet, filename: str) -> Optional[Tuple[int, int, int, int]]:
    """Optimized remarks finder with batch reading."""
    start_row, start_col = 2, 2
    try:
        search_end_col_idx = column_index_from_string(SEARCH_REMARKS_END_COL)
    except Exception:
        logger.error(f"[{filename}] Invalid column '{SEARCH_REMARKS_END_COL}' in config.")
        return None

    search_area = (start_row, SEARCH_REMARKS_END_ROW, start_col, search_end_col_idx)
    area_values = get_area_values_batch(sheet, search_area)

    for (row, col), value in area_values.items():
        if value and isinstance(value, str) and "REMARKS".lower() in value.lower():
            return start_row, row, start_col, col

    logger.error(f"[{filename}] Could not find 'REMARKS' cell in range B2:{SEARCH_REMARKS_END_COL}{SEARCH_REMARKS_END_ROW}")
    return None


def find_data_area_sheet1_optimized(sheet: Worksheet, end_col_from_first_area: int, filename: str, remarks_row: int) -> Optional[
    Tuple[int, int, int, int]]:
    """Modified data area finder for Sheet 1 - uses 'Approval Path' to find end row."""
    col_b_values = get_column_values_batch(sheet, 2, 1, sheet.max_row)

    start_row = None
    for i, cell_val in enumerate(col_b_values[:min(len(col_b_values), 99)]):
        if cell_val == 1:
            start_row = i + 1
            break

    if not start_row:
        logger.info(f"[{filename}] Could not find start row with value '1' in Column B for Sheet 1. Using remarks row + 1 as fallback.")
        start_row = remarks_row + 1

    end_row = None
    for i, cell_val in enumerate(col_b_values[start_row - 1:]):
        if isinstance(cell_val, str) and cell_val.strip() == "Approval Path":
            end_row = start_row + i
            break

    if not end_row:
        logger.error(f"[{filename}] Could not find 'Approval Path' in Column B.")
        return None

    return start_row, end_row, 2, end_col_from_first_area


def find_total_joint_from_data_area(sheet: Worksheet, start_row: int, end_row: int) -> Any:
    """Find the biggest number in Column B within the data area range."""
    max_number = None
    col_b_values = get_column_values_batch(sheet, 2, start_row, end_row)

    for cell_val in col_b_values:
        if isinstance(cell_val, (int, float)):
            if max_number is None or cell_val > max_number:
                max_number = cell_val

    return max_number if max_number is not None else "Not Found"


def find_areas_visual_optimized(sheet: Worksheet, report_type: str, filename: str) -> Optional[
    Tuple[Tuple[int, int, int, int], Tuple[int, int, int, int]]]:
    """Optimized visual area finder, using 'inspector' to determine the end row."""
    try:
        start_col_data = 2 if report_type == "v2" else 1
        remarks_col_idx = column_index_from_string("AA")

        data_start_row = None
        col_start_values = get_column_values_batch(sheet, start_col_data, 1, min(sheet.max_row, 100))
        for i, cell_val in enumerate(col_start_values):
            if cell_val == 1:
                data_start_row = i + 1
                break

        if not data_start_row:
            logger.info(f"[{filename}] Could not find start row with value '1' in Column {get_column_letter(start_col_data)} for Sheet 2. Searching for 'REMARKS' as fallback.")
            remarks_cell_row = None
            search_end_col_idx = column_index_from_string(SEARCH_REMARKS_END_COL)
            search_area = (2, SEARCH_REMARKS_END_ROW, 2, search_end_col_idx)
            area_values = get_area_values_batch(sheet, search_area)
            for (row, col), value in area_values.items():
                if value and isinstance(value, str) and "REMARKS".lower() in value.lower():
                    remarks_cell_row = row
                    break
            
            if remarks_cell_row:
                data_start_row = remarks_cell_row + 1
                logger.info(f"[{filename}] Found 'REMARKS' at row {remarks_cell_row}. Using row {data_start_row} as data start row.")
            else:
                logger.error(
                    f"[{filename}] Could not find data start row with value '1' and no 'REMARKS' cell found for fallback.")
                return None

        data_end_row = None
        col_all_values = get_column_values_batch(sheet, start_col_data, data_start_row, sheet.max_row)
        for i, cell_val in enumerate(col_all_values):
            if isinstance(cell_val, str) and "inspector" in cell_val.lower():
                data_end_row = data_start_row + i
                break

        if not data_end_row:
            logger.error(
                f"[{filename}] Could not find 'inspector' text in Column {get_column_letter(start_col_data)} to determine data end row.")
            return None

        header_area = (2, data_start_row - 1, 1, remarks_col_idx)
        data_area = (data_start_row, data_end_row, start_col_data, remarks_col_idx)
        return header_area, data_area

    except Exception as e:
        logger.error(f"[{filename}] Error in find_areas_visual_optimized: {e}")
        return None


def determine_document_version(workbook, filename: str, sheet1_name=None) -> str:
    """Determine the document version based on sheet names and structure."""
    try:
        sheet_names = workbook.sheetnames
        v1_indicators = [SHEET_1_NAME_V1 in sheet_names, SHEET_1_NAME_V1_ALT in sheet_names, SHEET_2_NAME_V1 in sheet_names]
        v2_indicators = [SHEET_1_NAME_V2 in sheet_names, SHEET_2_NAME_V2 in sheet_names]

        if any(v1_indicators):
            return "v1"
        elif any(v2_indicators):
            return "v2"
        else:
            return "unknown"

    except Exception as e:
        logger.warning(f"[{filename}] Error determining document version: {e}")
        return "unknown"


# --- Optimized Extraction Functions ---

def extract_min_thickness_and_joint_optimized(sheet: Worksheet, data_area: Tuple[int, int, int, int],
                                              header_area: Tuple[int, int, int, int], filename: str) -> Tuple[Any, str]:
    """Optimized minimum thickness extraction, excluding zero values."""
    if not header_area:
        return "Not Found", "Not Found"

    thickness_col = find_column_by_keyword_optimized(sheet, header_area, "MINIMUM THICKNESS", filename)
    if not thickness_col:
        return "Not Found", "Not Found"

    start_row, end_row, _, _ = data_area
    thickness_vals = get_column_values_batch(sheet, thickness_col, start_row, end_row)
    col_b_vals = get_column_values_batch(sheet, 2, start_row, end_row)
    col_c_vals = get_column_values_batch(sheet, 3, start_row, end_row)

    min_thickness_data = []
    for i, cell_val in enumerate(thickness_vals):
        if isinstance(cell_val, (int, float)) and cell_val > 0:
            min_thickness_data.append({'value': cell_val, 'index': i})

    if not min_thickness_data:
        return "Not Found", "Not Found"

    min_item = min(min_thickness_data, key=lambda x: x['value'])
    row_idx = min_item['index']
    val_b = col_b_vals[row_idx]
    val_c = col_c_vals[row_idx]
    joint = f"{val_b or ''}{val_c or ''}"

    return min_item['value'], joint


def extract_max_thickness_and_joint_optimized(sheet: Worksheet, data_area: Tuple[int, int, int, int],
                                              header_area: Tuple[int, int, int, int], filename: str) -> Tuple[Any, str]:
    """Optimized maximum thickness extraction."""
    if not header_area:
        return "Not Found", "Not Found"

    thickness_col = find_column_by_keyword_optimized(sheet, header_area, "MAXIMUM THICKNESS", filename)
    if not thickness_col:
        return "Not Found", "Not Found"

    start_row, end_row, _, _ = data_area
    thickness_vals = get_column_values_batch(sheet, thickness_col, start_row, end_row)
    col_b_vals = get_column_values_batch(sheet, 2, start_row, end_row)
    col_c_vals = get_column_values_batch(sheet, 3, start_row, end_row)

    thickness_data = []
    for i, cell_val in enumerate(thickness_vals):
        if isinstance(cell_val, (int, float)):
            thickness_data.append({'value': cell_val, 'index': i})

    if not thickness_data:
        return "Not Found", "Not Found"

    max_item = max(thickness_data, key=lambda x: x['value'])
    row_idx = max_item['index']
    val_b = col_b_vals[row_idx]
    val_c = col_c_vals[row_idx]
    joint = f"{val_b or ''}{val_c or ''}"

    return max_item['value'], joint


def extract_max_thickness_pipe_joint_optimized(sheet: Worksheet, data_area: Tuple[int, int, int, int],
                                               header_area: Tuple[int, int, int, int], filename: str) -> Any:
    """Extract maximum thickness only for joints where JOINT TYPES is 'PIPE'."""
    if not header_area:
        return "="

    thickness_col = find_column_by_keyword_optimized(sheet, header_area, "MAXIMUM THICKNESS", filename)
    joint_type_col = find_column_by_keyword_optimized(sheet, header_area, "JOINT TYPES", filename)
    if not thickness_col or not joint_type_col:
        return "="

    start_row, end_row, _, _ = data_area
    thickness_vals = get_column_values_batch(sheet, thickness_col, start_row, end_row)
    joint_type_vals = get_column_values_batch(sheet, joint_type_col, start_row, end_row)

    max_thickness = None
    for i, joint_type_val in enumerate(joint_type_vals):
        thickness_val = thickness_vals[i]
        if isinstance(joint_type_val, str) and joint_type_val.strip().upper() == "PIPE":
            if isinstance(thickness_val, (int, float)):
                if max_thickness is None or thickness_val > max_thickness:
                    max_thickness = thickness_val
    return max_thickness if max_thickness is not None else "="


def extract_from_sheet1_optimized(sheet: Worksheet, report_type: str, filename: str) -> Tuple[Dict[str, Any], Optional[Tuple[int, int, int, int]]]:
    """Optimized extraction function for Sheet 1."""
    logger.info(f"[{filename}] Processing sheet: {sheet.title} (Type: {report_type})")

    header_area = find_area_by_remarks_optimized(sheet, filename)
    if not header_area:
        return {}, None

    start_r, end_r, start_c, end_c = header_area
    logger.info(f"[{filename}] Header Area: {get_column_letter(start_c)}{start_r}:{get_column_letter(end_c)}{end_r}")

    data = {
        "Inspection Date Finish": sheet['E4'].value,
        "Length Of Inspection (m)": search_for_value_optimized(sheet, header_area, ["LENGTH"], offset=2),
        "Pipe Material": search_for_value_optimized(sheet, header_area, ["PIPE MATERIAL"], offset=2),
        "Operating Pressure (psi)": search_for_value_optimized(sheet, header_area, ["PRESSURE"], offset=2),
        "Operating Temperature (F)": search_for_value_optimized(sheet, header_area, ["TEMPERATURE"], offset=2),
        "Service Fluid": search_for_value_optimized(sheet, header_area, ["FLUID"], offset=2),
        "Line ID": search_for_value_optimized(sheet, header_area, ["LINE ID"], offset=2),
        "NPS (in)": search_for_value_optimized(sheet, header_area, ["NOMINAL PIPE SIZE"], offset=2),
        "Estimation Year Buil": search_for_value_optimized(sheet, header_area, ["YEAR"], offset=2),
        "Pipe Segment": search_for_value_optimized(sheet, header_area, ["PIPE SEGMENT"], offset=2)
    }

    if report_type == "v1":
        data["Length Of Inspection (m)"] = search_for_value_optimized(sheet, header_area, ["LENGTH"], offset=2)
        data["Nominal Thickness (in)"] = search_for_value_optimized(sheet, header_area, ["NOMINAL THICKNESS"],
                                                                    offset=2, row_offset=1)
        data["Flange Rating"] = find_value_in_cell_optimized(sheet, header_area, "ANSI")
    elif report_type == "v2":
        data["Length Of Inspection (m)"] = search_for_value_optimized(sheet, header_area, ["LENGTH"], offset=1)
        data["Nominal Thickness (mm)"] = search_for_value_optimized(sheet, header_area, ["NOMINAL THICKNESS"], offset=2)
        data["Flange Rating"] = search_for_value_optimized(sheet, header_area, ["ANSI"], row_offset=1)

    data_area = find_data_area_sheet1_optimized(sheet, end_c, filename, remarks_row=end_r)

    if data_area:
        logger.info(
            f"[{filename}] Data Area: {get_column_letter(data_area[2])}{data_area[0]}:{get_column_letter(data_area[3])}{data_area[1]}")

        start_row, end_row, _, _ = data_area
        total_joint_calculated = find_total_joint_from_data_area(sheet, start_row, end_row)
        data["Total Joint"] = total_joint_calculated

        header_area_2 = (end_r, data_area[0] - 1, 2, end_c)

        min_thick, min_joint_loc = extract_min_thickness_and_joint_optimized(sheet, data_area, header_area_2, filename)
        data["Minimum Thickness (mm)"] = min_thick
        data["Joint of Minimum Thickness"] = min_joint_loc

        max_thick, max_joint_loc = extract_max_thickness_and_joint_optimized(sheet, data_area, header_area_2, filename)
        data["Maximum Thickness (mm)"] = max_thick
        data["Joint of Maximum Thickness"] = max_joint_loc

        max_thick_pipe = extract_max_thickness_pipe_joint_optimized(sheet, data_area, header_area_2, filename)
        data["Max Thickness Pipe Joint"] = max_thick_pipe

        data["Remarks"] = extract_unique_values_from_col_optimized(sheet, data_area, end_c)

    return data, data_area


def calculate_inspection_rate_ratio(sheet, data_area, nps_size, filename: str):
    """Calculate inspection rate ratio for each joint in Sheet 1."""
    logger.info(f"[{filename}] Calculating inspection rate ratios...")

    try:
        nps_value = float(nps_size) if nps_size and nps_size != "Not Found" else 0
        if nps_value < 4:
            divisor = 2
        elif 4 <= nps_value <= 12:
            divisor = 4
        else:
            divisor = 6
        logger.info(f"[{filename}] NPS size: {nps_value}, using divisor: {divisor}")
    except (ValueError, TypeError):
        logger.warning(f"[{filename}] Invalid NPS size '{nps_size}', using default divisor of 4")
        divisor = 4

    inspection_start_col = 4
    inspection_end_col = 15
    joint_col = 2
    joint_ratios = {}
    start_row, end_row, _, _ = data_area

    data_block = list(sheet.iter_rows(min_row=start_row, max_row=end_row,
                                      min_col=joint_col, max_col=inspection_end_col,
                                      values_only=True))

    joint_col_idx = 0
    inspection_start_idx = inspection_start_col - joint_col
    inspection_end_idx = inspection_end_col - joint_col

    for row_data in data_block:
        joint_val = row_data[joint_col_idx]
        if joint_val is not None:
            try:
                joint_num = int(joint_val)
                inspection_values = row_data[inspection_start_idx: inspection_end_idx + 1]
                numeric_count = sum(
                    1 for cell_val in inspection_values if
                    cell_val is not None and cell_val != "" and isinstance(cell_val, (int, float)))
                ratio = numeric_count / divisor
                if joint_num not in joint_ratios:
                    joint_ratios[joint_num] = []
                joint_ratios[joint_num].append(ratio)
            except (ValueError, TypeError):
                logger.warning(f"[{filename}] Invalid joint value: {joint_val}")
                continue

    joint_avg_ratios = {}
    for joint_num, ratios in joint_ratios.items():
        if ratios:
            avg_ratio = sum(ratios) / len(ratios)
            joint_avg_ratios[joint_num] = round(avg_ratio, 3)

    return joint_avg_ratios


def write_minimum_thickness_to_sheet2(workbook, sheet1, sheet2, sheet1_data_area, sheet2_data_area,
                                      sheet1_header_area, report_type, filename: str):
    """Write minimum thickness values from Sheet 1 to Sheet 2."""
    logger.info(f"[{filename}] Writing minimum thickness values to Sheet 2...")

    sheet1_data_header_area = (sheet1_header_area[1], sheet1_data_area[0] - 1, 2, sheet1_header_area[3])
    thickness_col = find_column_by_keyword_optimized(sheet1, sheet1_data_header_area, "MINIMUM THICKNESS", filename)
    if not thickness_col:
        logger.error(f"[{filename}] Could not find thickness column in Sheet 1")
        return False

    sheet1_joint_col = 2
    s1_start_row, s1_end_row, _, _ = sheet1_data_area
    joint_vals = get_column_values_batch(sheet1, sheet1_joint_col, s1_start_row, s1_end_row)
    thickness_vals = get_column_values_batch(sheet1, thickness_col, s1_start_row, s1_end_row)

    joint_thickness_map = {}
    for i, joint_val in enumerate(joint_vals):
        thickness_val = thickness_vals[i]
        if joint_val is not None and thickness_val is not None:
            try:
                joint_num = int(joint_val)
                thickness_num = float(thickness_val)
                if joint_num not in joint_thickness_map:
                    joint_thickness_map[joint_num] = []
                joint_thickness_map[joint_num].append(thickness_num)
            except (ValueError, TypeError):
                continue

    joint_min_thickness = {}
    for joint_num, thickness_list in joint_thickness_map.items():
        if thickness_list:
            non_zero_values = [t for t in thickness_list if t > 0]
            if non_zero_values:
                joint_min_thickness[joint_num] = min(non_zero_values)
            else:
                joint_min_thickness[joint_num] = 0

    sheet2_joint_col = 1 if report_type == "v1" else 2
    s2_start_row, s2_end_row, _, _ = sheet2_data_area
    sheet2_joint_vals = get_column_values_batch(sheet2, sheet2_joint_col, s2_start_row, s2_end_row)

    write_col = sheet2_data_area[3] + 1
    header_row = sheet2_data_area[0] - 1
    sheet2.cell(row=header_row, column=write_col, value="Min Thickness")

    written_count = 0
    for i, joint_val in enumerate(sheet2_joint_vals):
        row = s2_start_row + i
        if joint_val is not None:
            try:
                joint_num = int(joint_val)
                if joint_num in joint_min_thickness:
                    sheet2.cell(row=row, column=write_col, value=joint_min_thickness[joint_num])
                    written_count += 1
                else:
                    sheet2.cell(row=row, column=write_col, value="N/A")
            except (ValueError, TypeError):
                sheet2.cell(row=row, column=write_col, value="N/A")

    logger.info(f"[{filename}] Successfully wrote {written_count} minimum thickness values to Sheet 2")
    return True


def write_inspection_rate_ratio_to_sheet2(workbook, sheet1, sheet2, sheet1_data_area, sheet2_data_area,
                                          sheet1_header_area, report_type, filename: str):
    """Calculate inspection rate ratios from Sheet 1 and write them to Sheet 2."""
    logger.info(f"[{filename}] Writing inspection rate ratios to Sheet 2...")

    sheet1_header_search_area = find_area_by_remarks_optimized(sheet1, filename)
    if sheet1_header_search_area:
        nps_size = search_for_value_optimized(sheet1, sheet1_header_search_area, ["NOMINAL PIPE SIZE"], offset=2)
    else:
        nps_size = "Not Found"
        logger.warning(f"[{filename}] Could not find header area to get NPS size")

    joint_ratios = calculate_inspection_rate_ratio(sheet1, sheet1_data_area, nps_size, filename)

    if not joint_ratios:
        logger.warning(f"[{filename}] No inspection rate ratios calculated")
        return False

    sheet2_joint_col = 1 if report_type == "v1" else 2
    s2_start_row, s2_end_row, _, _ = sheet2_data_area
    sheet2_joint_vals = get_column_values_batch(sheet2, sheet2_joint_col, s2_start_row, s2_end_row)

    ratio_col = sheet2_data_area[3] + 2
    header_row = sheet2_data_area[0] - 1
    sheet2.cell(row=header_row, column=ratio_col, value="Inspection Rate Ratio")

    written_count = 0
    for i, joint_val in enumerate(sheet2_joint_vals):
        row = s2_start_row + i
        if joint_val is not None:
            try:
                joint_num = int(joint_val)
                if joint_num in joint_ratios:
                    sheet2.cell(row=row, column=ratio_col, value=joint_ratios[joint_num])
                    written_count += 1
                else:
                    sheet2.cell(row=row, column=ratio_col, value="N/A")
            except (ValueError, TypeError):
                sheet2.cell(row=row, column=ratio_col, value="N/A")

    logger.info(f"[{filename}] Successfully wrote {written_count} inspection rate ratios to Sheet 2")
    return True


def extract_from_sheet2_optimized(sheet: Worksheet, report_type: str, filename: str) -> Dict[str, Any]:
    """Optimized extraction function for Sheet 2."""
    logger.info(f"[{filename}] Processing sheet: {sheet.title} (Type: {report_type})")

    areas = find_areas_visual_optimized(sheet, report_type, filename)
    if not areas:
        return {}

    header_area, data_area = areas
    logger.info(
        f"[{filename}] Header Area: {get_column_letter(header_area[2])}{header_area[0]}:{get_column_letter(header_area[3])}{header_area[1]}")
    logger.info(
        f"[{filename}] Data Area: {get_column_letter(data_area[2])}{data_area[0]}:{get_column_letter(data_area[3])}{data_area[1]}")

    condition_keywords = ["Above Ground", "Lay Down", "Under Ground", "Sleeve", "Clamp", "Isolation"]
    data = {}

    for keyword in condition_keywords:
        col_idx = find_column_by_keyword_optimized(sheet, header_area, keyword, filename)
        unique_vals = extract_unique_values_from_col_or_notfound(sheet, data_area, col_idx)
        data[f"{keyword} Position" if 'Ground' in keyword or 'Lay Down' in keyword else f"{keyword} Condition"] = unique_vals

    painting_col_idx = find_column_by_keyword_optimized(sheet, header_area, "Painting", filename)
    data["Painting Condition"] = extract_unique_values_from_col_optimized(sheet, data_area, painting_col_idx)

    remarks_col_idx = header_area[3]
    data["Remarks Visual"] = extract_unique_values_from_col_optimized(sheet, data_area, remarks_col_idx)

    return data


def process_single_file_extraction_only(file_path: str) -> Dict[str, Any]:
    """Processes a single Excel file to extract data without modifying the original file."""
    filename = os.path.basename(file_path)
    logger.info(f"[{filename}] Extracting data from file")

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        combined_data = {"File Name": filename}
        document_version = determine_document_version(workbook, filename)
        combined_data["Document Version"] = document_version

        if document_version == "unknown":
            logger.warning(f"[{filename}] has an unknown template. Recording with version only.")
            workbook.close()
            return combined_data

        report_type = document_version
        sheet1, sheet2 = None, None

        if report_type == "v1":
            sheet1 = workbook[SHEET_1_NAME_V1] if SHEET_1_NAME_V1 in workbook.sheetnames else workbook[SHEET_1_NAME_V1_ALT]
            if SHEET_2_NAME_V1 in workbook.sheetnames:
                sheet2 = workbook[SHEET_2_NAME_V1]
        elif report_type == "v2":
            sheet1 = workbook[SHEET_1_NAME_V2]
            if SHEET_2_NAME_V2 in workbook.sheetnames:
                sheet2 = workbook[SHEET_2_NAME_V2]

        if not sheet1:
            logger.warning(f"[{filename}] Could not find primary data sheet (Version: {report_type}). Treating as unknown.")
            workbook.close()
            return {"File Name": filename, "Document Version": document_version}

        data_from_s1, _ = extract_from_sheet1_optimized(sheet1, report_type, filename)
        combined_data.update(data_from_s1)

        if sheet2:
            data_from_s2 = extract_from_sheet2_optimized(sheet2, report_type, filename)
            combined_data.update(data_from_s2)
        else:
            logger.warning(f"[{filename}] Sheet 2 not found, skipping visual data extraction.")

        workbook.close()
        return combined_data

    except Exception as e:
        logger.error(f"[{filename}] Error processing for extraction only: {str(e)}")
        return {"File Name": filename, "Error": str(e)}


def process_single_file_with_calculations_and_updates(file_path: str) -> Dict[str, Any]:
    """Processes a single Excel file, performs calculations, and writes results back to the original file."""
    filename = os.path.basename(file_path)
    logger.info(f"[{filename}] Processing file for calculations and updates")

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
        document_version = determine_document_version(workbook, filename)

        if document_version == "unknown":
            logger.warning(f"[{filename}] has an unknown template. Skipping calculations and updates.")
            workbook.close()
            return {"File Name": filename, "Error": "Unknown template; skipped update."}

        combined_data = {"File Name": filename}
        report_type = document_version
        sheet1, sheet2 = None, None

        if report_type == "v1":
            sheet1 = workbook[SHEET_1_NAME_V1] if SHEET_1_NAME_V1 in workbook.sheetnames else workbook[SHEET_1_NAME_V1_ALT]
            if SHEET_2_NAME_V1 in workbook.sheetnames:
                sheet2 = workbook[SHEET_2_NAME_V1]
        elif report_type == "v2":
            sheet1 = workbook[SHEET_1_NAME_V2]
            if SHEET_2_NAME_V2 in workbook.sheetnames:
                sheet2 = workbook[SHEET_2_NAME_V2]

        if not sheet1:
            logger.warning(f"[{filename}] Could not find primary data sheet (Version: {report_type}). Skipping calculations.")
            workbook.close()
            return {"File Name": filename, "Error": "Could not find primary data sheet"}

        sheet1_header_area = find_area_by_remarks_optimized(sheet1, filename)
        if not sheet1_header_area:
            logger.error(f"[{filename}] Could not find header area in Sheet 1")
            workbook.close()
            return {"File Name": filename, "Error": "Could not find header area in Sheet 1"}

        data_from_s1, sheet1_data_area = extract_from_sheet1_optimized(sheet1, report_type, filename)
        combined_data.update(data_from_s1)

        if sheet2 and sheet1_data_area:
            sheet2_areas = find_areas_visual_optimized(sheet2, report_type, filename)
            if sheet2_areas:
                _, sheet2_data_area = sheet2_areas

                thickness_success = write_minimum_thickness_to_sheet2(
                    workbook, sheet1, sheet2, sheet1_data_area, sheet2_data_area, sheet1_header_area, report_type,
                    filename
                )

                ratio_success = write_inspection_rate_ratio_to_sheet2(
                    workbook, sheet1, sheet2, sheet1_data_area, sheet2_data_area, sheet1_header_area, report_type,
                    filename
                )

                if thickness_success or ratio_success:
                    workbook.save(file_path)
                    logger.info(f"[{filename}] Successfully updated with calculated values.")
                else:
                    logger.warning(f"[{filename}] No calculations were successfully written.")

            else:
                logger.warning(f"[{filename}] Could not find Sheet 2 areas, skipping calculations.")
        else:
            if not sheet2:
                logger.warning(f"[{filename}] Sheet 2 not found, skipping calculations.")
            if not sheet1_data_area:
                logger.warning(f"[{filename}] Sheet 1 data area not found, skipping calculations.")

        workbook.close()
        return combined_data

    except Exception as e:
        logger.error(f"[{filename}] Error processing for calculations and updates: {str(e)}")
        return {"File Name": filename, "Error": str(e)}


def process_files_parallel(file_paths: List[str], process_func) -> List[Dict[str, Any]]:
    """Process multiple Excel files in parallel using a given processing function."""
    max_workers = min(len(file_paths), MAX_WORKERS)
    results = []

    logger.info(f"Processing {len(file_paths)} files using {max_workers} workers")

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        future_to_file = {executor.submit(process_func, fp): fp for fp in file_paths}

        for future in as_completed(future_to_file):
            file_path = future_to_file[future]
            filename = os.path.basename(file_path)
            try:
                result = future.result()
                results.append(result)
                logger.info(f"[{filename}] Completed processing")
            except Exception as e:
                logger.error(f"[{filename}] Error processing file: {e}")
                results.append({"File Name": filename, "Error": str(e)})

    return results


# --- Main Orchestration Functions ---

def run_extraction_only():
    """Main function for extraction only. Processes files and generates a summary Excel file."""
    start_time = time.time()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, DATA_FOLDER)
    output_dir = os.path.join(script_dir, OUTPUT_FOLDER)

    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isdir(data_dir):
        logger.error(f"The directory '{DATA_FOLDER}' was not found.")
        return

    excel_files = [os.path.join(data_dir, f) for f in os.listdir(data_dir) if
                   f.endswith(".xlsx") and not f.startswith("~")]

    if not excel_files:
        logger.warning("No valid Excel files (.xlsx) found in the data directory.")
        return

    logger.info(f"Found {len(excel_files)} Excel files to process for extraction only.")
    all_results = process_files_parallel(excel_files, process_single_file_extraction_only)

    valid_results = [r for r in all_results if "Error" not in r]
    error_results = [r for r in all_results if "Error" in r]

    if error_results:
        logger.warning(f"{len(error_results)} files had errors during extraction:")
        for error_result in error_results:
            logger.warning(f"  - {error_result['File Name']}: {error_result.get('Error', 'Unknown error')}")

    if valid_results:
        logger.info("Creating output Excel file with extracted data...")
        headers = [
            "File Name", "Document Version", "Inspection Date Finish", "Length Of Inspection (m)",
            "Pipe Material", "Nominal Thickness (in)", "Nominal Thickness (mm)",
            "Operating Pressure (psi)", "Operating Temperature (F)", "Service Fluid",
            "Flange Rating", "Line ID", "NPS (in)", "Total Joint",
            "Minimum Thickness (mm)", "Joint of Minimum Thickness", "Maximum Thickness (mm)",
            "Joint of Maximum Thickness", "Max Thickness Pipe Joint", "Remarks",
            "Above Ground Position", "Lay Down Position", "Under Ground Position",
            "Painting Condition", "Sleeve Condition", "Clamp Condition",
            "Isolation Condition", "Remarks Visual", "Estimation Year Buil", "Pipe Segment"
        ]

        output_workbook = openpyxl.Workbook()
        sheet = output_workbook.active
        sheet.title = "Extraction Summary"
        sheet.append(headers)

        for result in valid_results:
            row_data = [result.get(header, "") for header in headers]
            sheet.append(row_data)

        output_path = os.path.join(output_dir, "Extraction_Summary.xlsx")
        try:
            output_workbook.save(output_path)
            end_time = time.time()
            logger.info(
                f"Successfully extracted data from {len(valid_results)} files in {end_time - start_time:.2f} seconds")
            logger.info(f"Extraction summary saved to: {output_path}")
        except Exception as e:
            logger.error(f"Error saving extraction summary Excel file: {e}")
    else:
        logger.error("No valid data extracted from any files.")


def run_calculations_and_updates():
    """Main function for calculations and updates. Processes files and writes results back to them."""
    start_time = time.time()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, DATA_FOLDER)

    if not os.path.isdir(data_dir):
        logger.error(f"The directory '{DATA_FOLDER}' was not found.")
        return

    excel_files = [os.path.join(data_dir, f) for f in os.listdir(data_dir) if
                   f.endswith(".xlsx") and not f.startswith("~")]

    if not excel_files:
        logger.warning("No valid Excel files (.xlsx) found in the data directory.")
        return

    logger.info(f"Found {len(excel_files)} Excel files to process for calculations and updates.")
    all_results = process_files_parallel(excel_files, process_single_file_with_calculations_and_updates)

    successful_updates = [r for r in all_results if "Error" not in r]
    failed_updates = [r for r in all_results if "Error" in r]

    if failed_updates:
        logger.warning(f"{len(failed_updates)} files had errors during calculation/update:")
        for error_result in failed_updates:
            logger.warning(f"  - {error_result['File Name']}: {error_result.get('Error', 'Unknown error')}")

    end_time = time.time()
    logger.info(
        f"Finished processing {len(successful_updates)} files for calculations and updates in {end_time - start_time:.2f} seconds.")
    logger.info(f"Original files in '{DATA_FOLDER}' have been updated.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Excel Data Extractor and Calculator.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        '--mode',
        choices=['extract', 'calculate_and_update'],
        default='extract',
        help="""Choose the operation mode:
        'extract' (default): Extracts data from Excel files and creates a summary. Does NOT modify original files.
        'calculate_and_update': Calculates and writes results back to the original Excel files.
        """
    )

    args = parser.parse_args()

    if args.mode == 'extract':
        logger.info("Running in 'extraction only' mode.")
        run_extraction_only()
    elif args.mode == 'calculate_and_update':
        logger.info("Running in 'calculate and update' mode. Original files will be modified.")
        run_calculations_and_updates()